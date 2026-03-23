import openpyxl
import shutil

src = r"C:\Users\Administrator\Desktop\ceshi\temp_converted.xlsx"
dst = r"C:\Users\Administrator\Desktop\ceshi\深户团购报价清单_涨价15%.xlsx"

# 复制原文件
shutil.copy2(src, dst)

# 打开并检查图片
wb = openpyxl.load_workbook(dst)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# 检查是否有图片
print(f"\n图片数量: {len(ws._images)}")

# 打印关键报价单元格
print("\n原始报价（行2起）:")
for row in range(2, ws.max_row + 1):
    model = ws.cell(row, 2).value  # B列=型号
    unit_price = ws.cell(row, 6).value  # F列=单价
    total_price = ws.cell(row, 7).value  # G列=总价
    if model:
        print(f"  {model}: 单价={unit_price}, 总价={total_price}")

# 提升15%
print("\n正在提升15%...")
changes = 0
for row in range(2, ws.max_row + 1):
    unit_price = ws.cell(row, 6).value
    total_price = ws.cell(row, 7).value
    if isinstance(unit_price, (int, float)) and unit_price > 0:
        ws.cell(row, 6).value = round(unit_price * 1.15, 2)
        if isinstance(total_price, (int, float)) and total_price > 0:
            ws.cell(row, 7).value = round(total_price * 1.15, 2)
        changes += 1

print(f"已更新 {changes} 个单元格")

# 保存
wb.save(dst)
print(f"\n已保存: {dst}")

# 验证
wb2 = openpyxl.load_workbook(dst)
ws2 = wb2.active
print("\n涨价后报价:")
for row in range(2, ws2.max_row + 1):
    model = ws2.cell(row, 2).value
    unit_price = ws2.cell(row, 6).value
    total_price = ws2.cell(row, 7).value
    if model and isinstance(unit_price, (int, float)):
        print(f"  {model}: 单价={unit_price}, 总价={total_price}")
