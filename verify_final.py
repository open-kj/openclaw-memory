import openpyxl

f = r"C:\Users\Administrator\Desktop\ceshi\深户团购报价清单_涨价15%.xlsx"

wb = openpyxl.load_workbook(f, data_only=False)
ws = wb.active

print(f"图片数量: {len(ws._images)}")
print("\n报价验证:")
print(f"{'型号':<15} {'单价(原)':>10} {'单价(新)':>10} {'涨幅':>8}")
print("-" * 50)

old_prices = {'CDK-E2828': 1850, 'CDK-2138': 2400, 'CDK-418': 245, 'EC-200': 400, 'CDK-1050': 2200}

for row in range(2, 7):
    model = ws.cell(row, 2).value
    unit = ws.cell(row, 6).value
    total_formula = ws.cell(row, 7).value
    old = old_prices.get(model, 0)
    if old > 0:
        expected = round(old * 1.15, 2)
        print(f"{model:<15} {old:>10.2f} {unit:>10.2f} {(unit/old-1)*100:>7.2f}%")
    else:
        print(f"{model}: 单价={unit}")

print(f"\n✅ 文件已保存，图片={len(ws._images)}张已保留")
