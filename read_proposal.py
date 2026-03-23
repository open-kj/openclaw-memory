import openpyxl

f = r"C:\Users\Administrator\Desktop\ceshi\深户团购报价清单_涨价15%.xlsx"

wb = openpyxl.load_workbook(f, data_only=False)
ws = wb.active

print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print(f"图片数量: {len(ws._images)}")

print("\n=== 完整产品明细 ===")
for row in range(1, ws.max_row + 1):
    row_data = []
    for col in range(1, min(ws.max_column + 1, 20)):
        val = ws.cell(row, col).value
        if val is not None:
            row_data.append(f"Col{col}={repr(val)[:50]}")
    if row_data:
        print(f"\nRow {row}:")
        for d in row_data:
            print(f"  {d}")
